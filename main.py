import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import IconSet, FormatObject
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import IconSetRule

import sys, win32com.client
from win32com import *
from win32api import *
from win32com.client import *

from pyrfc import Connection
from pyrfc import ABAPApplicationError, ABAPRuntimeError, LogonError, CommunicationError, RFCError

from datetime import datetime

import os
import shutil

from Email import Email


import logging

import configparser

import socket

from pythonping import ping

from reportClient import ReportClient

from scrapingWeb import ScrapingWeb

from configuration import Config

config = Config()


isdirlog = os.path.exists('logs')
if not isdirlog:
    os.mkdir('logs')

dateTimeObj = datetime.now()
log_name = "logs\{}{}{}{}{}{}.log".format(dateTimeObj.year, dateTimeObj.month,dateTimeObj.day, dateTimeObj.hour, dateTimeObj.minute, dateTimeObj.second)

logging.basicConfig(filename=log_name, format='%(asctime)s:%(levelname)s:%(message)s', datefmt='%d/%m/%Y %I:%M:%S %p', level=logging.INFO)

# Open Workbook with template
wb = openpyxl.load_workbook(config.source)

sheet = wb[config.sheet_connections]
dashboard = wb[config.sheet_dashboard]


logging.info('Se leyo el archivo template y de configuración.')

isdir = os.path.exists(config.path_gen)
if not isdir:
    os.mkdir(config.path_gen)


writerow = int(config.ini_cell)
cuenta_anterior = ""
system_client = []


countFail = 0
countOK = 0
cell_status = ""

def is_valid_ip(address):
    try:
        socket.inet_aton(address)
    except socket.error:
        return False
    return True

def server_response(address):
    #response = ping(address, count=2)
    server_ok = True
    #for re in response:
    #    server_ok = re.success
    return server_ok

def get_filter(type):
    if type == "ABAP" or type == "JAVA":
        return True
    return False

def get_system(currentSheet, row):
    cell_user = "{}{}".format(config.config_user_col, row)
    cell_password = "{}{}".format(config.config_password_col, row)
    cell_server = "{}{}".format(config.config_server_col, row)
    cell_sysnr = "{}{}".format(config.config_sysnr_col, row)
    cell_client = "{}{}".format(config.config_client_col, row)

    cell_cuenta = "{}{}".format(config.config_account_col, row)
    cell_name = "{}{}".format(config.config_name_col, row)
    cell_instance = "{}{}".format(config.config_instance_col, row)
    cell_saprouter = "{}{}".format(config.config_saprouter_col, row)

    cell_systemid = "{}{}".format(config.config_systemid_col, row)
    cell_system = "{}{}".format(config.config_system_col, row)
    cell_enviroment = "{}{}".format(config.config_enviroment_col, row)
    cell_type = "{}{}".format(config.config_type_col, row)
    cell_version_portal = "{}{}".format(config.config_version_portal_col, row)

    ###
    print("cell {} con valor {}-{}-{}".format(cell_name, currentSheet[cell_name].value,
                                              currentSheet[cell_client].value,
                                              currentSheet[cell_user].value))
    logging.info("cell {} con valor {}-{}-{}".format(cell_name, currentSheet[cell_name].value,
                                                     currentSheet[cell_client].value,
                                                     currentSheet[cell_user].value))

    system = {
        'cuenta': str(currentSheet[cell_cuenta].value).strip(),  # A
        'name': str(currentSheet[cell_name].value).strip(),  # B
        'instance': currentSheet[cell_instance].value,  # C
        'user': currentSheet[cell_user].value,  # H
        'passwd': currentSheet[cell_password].value,  # L
        'ashost': currentSheet[cell_server].value,  # I
        'sysnr': currentSheet[cell_sysnr].value,  # J
        'client': currentSheet[cell_client].value,  # N

        'systemid': currentSheet[cell_systemid].value,  # C
        'system': currentSheet[cell_system].value,  # D
        'enviroment': currentSheet[cell_enviroment].value,  # E
        'type': currentSheet[cell_type].value,  # G
        'version': currentSheet[cell_version_portal].value,  # P,
        'connection': '',
    }
    return system, cell_user, cell_saprouter, cell_password, cell_server, cell_sysnr, cell_client

def find_specific_row(currentSheet, dashboard, specific_column):
    global config
   # print("currentSheet.max_row->{}".format(currentSheet.max_row))
    for row in range(1, currentSheet.max_row + 2):
  #      print("row->{}".format(row))
        for column in specific_column:
  #          print("column->{}".format(column))
            cell_name = "{}{}".format(column, row)
            print(currentSheet[cell_name].value)
            logging.info(currentSheet[cell_name].value)
            if currentSheet[cell_name].value is None or currentSheet[cell_name].value == "":
                ##############################
                # SE HACE EL LLAMADO A LA CLASE QUE PINTA
 #               print("VA A PNTAR 1")
                report = ReportClient(system_client, system['cuenta'], config.path_gen)
                report.build_report()
                ##############################
                return
            if get_filter(currentSheet[cell_name].value):
                system, cell_user, cell_saprouter, cell_password, cell_server, cell_sysnr, cell_client  = get_system(currentSheet, row)

                #params = get_connection_params()
                if currentSheet[cell_saprouter].value is None or str(currentSheet[cell_saprouter].value) == "":
                    params = {'user': str(currentSheet[cell_user].value),
                              'passwd': str(currentSheet[cell_password].value),
                              'ashost': str(currentSheet[cell_server].value),
                              'sysnr': str(currentSheet[cell_sysnr].value),
                              'client': str(currentSheet[cell_client].value)
                    }
                else:
                    params = {'user': str(currentSheet[cell_user].value),
                              'passwd': str(currentSheet[cell_password].value),
                              'ashost': str(currentSheet[cell_server].value),
                              'sysnr': str(currentSheet[cell_sysnr].value),
                              'client': str(currentSheet[cell_client].value),
                              'saprouter': str(currentSheet[cell_saprouter].value)}

                connect = True
                if currentSheet[cell_name].value == "ABAP"  and not(is_valid_ip(str(currentSheet[cell_server].value))):
                    connect = False
                    write_result(dashboard, system, "La dirección IP no es validad, favor revisar configuración.")
                    system['connection'] = ''

#                if currentSheet[cell_name].value == "ABAP" and currentSheet[cell_saprouter].value is None or str(currentSheet[cell_saprouter].value) == "":# and server_response(str(currentSheet[cell_server].value)):
#                    connect = False
#                    write_result(dashboard, system, "No se tiene acceso a la dirección IP, revisar que se tenga acceso a ella.")
#                    system['connection'] = ''

                if currentSheet[cell_user].value is None or currentSheet[cell_user].value == "":
                    connect = False
                    write_result(dashboard, system, "No existe usuario en la configuración.")
                    system['connection'] = ''

                if currentSheet[cell_name].value == "JAVA" and (system["version"] is None or system["version"] == ""):
                    connect = False
                    write_result(dashboard, system, "No existe versión en la configuración para portales.")
                    system['connection'] = ''

                if connect:
#                    print("SYSTEM->{}".format(system))
                    if currentSheet[cell_name].value == "ABAP":
                        result = connect_rfc(**params)
                    elif currentSheet[cell_name].value == "JAVA":
                        # Diferentes versiones de portales soportadas hasta el momento
                        # NW 7.0
                        # Fiori 7.5
                        # BO
                        scraping = ScrapingWeb(system["version"], system["ashost"], system["user"], system["passwd"])
                        result = scraping.test_connection()

                    if result == "Connection OK":
                        logging.info(result)
                    else:
                        logging.error(result)
                    print(result)
                    write_result(dashboard, system, result)

                system_client.append(system)

    return row

def get_error(ex):
    error = {}
    ex_type_full = str(type(ex))
    error["type"] = ex_type_full[ex_type_full.rfind(".") + 1 : ex_type_full.rfind("'")]
    error["code"] = ex.code if hasattr(ex, "code") else "<None>"
    error["key"] = ex.key if hasattr(ex, "key") else "<None>"
    error["message"] = ex.message.split("\n")
    error["msg_class"] = ex.msg_class if hasattr(ex, "msg_class") else "<None>"
    error["msg_type"] = ex.msg_type if hasattr(ex, "msg_type") else "<None>"
    error["msg_number"] = ex.msg_number if hasattr(ex, "msg_number") else "<None>"
    error["msg_v1"] = ex.msg_v1 if hasattr(ex, "msg_v1") else "<None>"
    return error

def connect_rfc(**params):
    result = ""
    try:
        with Connection(**params) as conn:
            #result = conn.call('STFC_CONNECTION', REQUTEXT=u'Hello SAP!')
            result = "Connection OK"
            conn.close()
    except CommunicationError as err:
        err1 = get_error(err)
        #result = 'CommunicationError '.join(err1["message"])
        host = 'N/A'
        for key, value in params.items():
            if key == 'ashost': #antes era is y no ==
                host  = value
        message = 'Error de comunicación, se alcanzó el tiempo límite de espera, no se pudo alcanzar el servidor. ' + host
        result = message +" "+ str(err1["message"])
    except LogonError as err:
        err1 = get_error(err)
        result = ' '.join(err1["message"])
    except (ABAPApplicationError):
        result = "Ocurrio un error ABAPApplicationError."
    except ABAPRuntimeError as err:
        err1 = get_error(err)
        result = ' '.join(err1["message"])
        #result = "Ocurrio un error ABAPRuntimeError."
    except RFCError:
        result = "An error occurred."
    return result

def create_format(dashboard):
    global countFail, countOK, cell_status
    promedio = (countOK * 100) / (countOK + countFail)

    first = FormatObject(type='num', val=config.conf_threshold_first)
    second = FormatObject(type='num', val=config.conf_threshold_second)
    third = FormatObject(type='num', val=config.conf_threshold_third)
    iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third])#, showValue=None, percent=None,
                      #reverse=None)

    # aisgnación de regla
    dashboard[cell_status].value = promedio
    color_scale_rule = Rule(type='iconSet', iconSet=iconset)

    dashboard.conditional_formatting.add(cell_status, color_scale_rule)

def create_cell(column, row):
    return "{}{}".format(column, row)

def insert_date_cell(sheet,column, row, data):
    cell = create_cell(column, row)
    sheet[cell].value = data
    return cell

def write_result(dashboard, system, result ):
    global cuenta_anterior, writerow, countFail, countOK, cell_status, system_client, config
    #print("VA A PNTAR write_result")
    if(cuenta_anterior != system["cuenta"]):
     #   print("cuenta_anterior != system[cuenta]->{} != {}".format(cuenta_anterior, system["cuenta"]))
        if( countFail >= 1 or countOK >= 1):
            create_format(dashboard)
            ##############################
            # SE HACE EL LLAMADO A LA CLASE QUE PINTA
            report = ReportClient(system_client, cuenta_anterior, config.path_gen)
            report.build_report()
            ##############################

            system_client = []

        create_header(system["cuenta"])
        countFail = 0
        countOK = 0

    cuenta_anterior = system["cuenta"]
    writerow = writerow + 1

    cell = insert_date_cell(dashboard, "B", writerow, system["name"])
    cell = insert_date_cell(dashboard, "C", writerow, system["instance"])
    cell = insert_date_cell(dashboard, "D", writerow, system["ashost"])

    dateTimeObj = datetime.now()
    hora = "{}{}{}{}{}".format(dateTimeObj.hour, ':', dateTimeObj.minute, ':', dateTimeObj.second)

    cell = insert_date_cell(dashboard, "E", writerow, hora)

    cell = create_cell("G", writerow)
    if (result == "Connection OK"):
        countOK = countOK + 1
        dashboard[cell].fill = PatternFill(fgColor="548235", fill_type="solid")
        system['connection'] = 'OK'
    else:
        countFail = countFail + 1
        dashboard[cell].fill = PatternFill(fgColor="ff0000", fill_type="solid")

    cell = insert_date_cell(dashboard, "H", writerow, result)

    #####
    #Adding date, client, system type
    day = "{}{}{}{}{}".format(dateTimeObj.day, '/', dateTimeObj.month, '/', dateTimeObj.year)
    cell = insert_date_cell(dashboard, "I", writerow, system["cuenta"])
    cell = insert_date_cell(dashboard, "J", writerow, system["type"])
    cell = insert_date_cell(dashboard, "K", writerow, day)    

    #####

    if (countFail >= 1 or countOK >= 1):
        create_format(dashboard)


def create_header(cuenta):
    global writerow, cell_status

    writerow = writerow + 4

    #cell_header1_B = insert_date_cell(dashboard, "B", writerow,"System ERP")
    cell_header1_B = insert_date_cell(dashboard, "B", writerow, cuenta)
    cell_header1_F = insert_date_cell(dashboard, "F", writerow, "General Status:")

    cell_header1_D = create_cell("D", writerow)
    cell_header1_E = create_cell("E", writerow)
    cell_header1_G = create_cell("G", writerow)

    cell_status = create_cell("G", writerow)

    cell_header1_C = create_cell("C", writerow)

    font = Font(color=colors.BLACK, bold=True)
    cell_design_header1 = PatternFill(fgColor="8EA9DB", fill_type="solid")

    dashboard[cell_header1_B].fill = cell_design_header1
    dashboard[cell_header1_B].font = font

    dashboard[cell_header1_C].font = font
    dashboard[cell_header1_C].fill = cell_design_header1

    dashboard[cell_header1_F].font = font
    dashboard[cell_header1_F].fill = cell_design_header1

    dashboard[cell_header1_D].fill = cell_design_header1
    dashboard[cell_header1_E].fill = cell_design_header1
    dashboard[cell_header1_G].fill = cell_design_header1

    writerow = writerow + 1

    cell_design_header2 = PatternFill(fgColor="92D050", fill_type="solid")

    cell_header2_B = insert_date_cell(dashboard, "B", writerow, "Name")
    cell_header2_C = insert_date_cell(dashboard, "C", writerow, "Instancia")
    cell_header2_D = insert_date_cell(dashboard, "D", writerow, "Server")
    cell_header2_E = insert_date_cell(dashboard, "E", writerow, "Hora")

    dashboard[cell_header2_B].fill = cell_design_header2
    dashboard[cell_header2_C].fill = cell_design_header2
    dashboard[cell_header2_D].fill = cell_design_header2
    dashboard[cell_header2_E].fill = cell_design_header2

    cell_header2_E = create_cell("E", writerow)
    dashboard[cell_header2_E].fill = cell_design_header2
    cell_header2_F = create_cell("F", writerow)
    dashboard[cell_header2_F].fill = cell_design_header2
    cell_header2_G = create_cell("G", writerow)
    dashboard[cell_header2_G].fill = cell_design_header2


def backup_files_gen():
    dateTimeObj = datetime.now()
    folder = "{}{}{}_{}".format(dateTimeObj.year, dateTimeObj.month, dateTimeObj.day, dateTimeObj.hour)#, dateTimeObj.minute)
    files = os.listdir(config.path_gen)

    for file in files:
        if ".pdf" in file or ".xlsx" in file:
            isdirbackup = os.path.exists(os.path.join(config.path_gen, folder))
            if not isdirbackup:
                print("Directorio para Backup: {}".format(os.path.join(config.path_gen, folder)))
                logging.info("Directorio para Backup: {}".format(os.path.join(config.path_gen, folder)))

                os.mkdir(config.path_gen + folder)
            print("Se envia a backup el archivo: {}".format(file))
            logging.info("Se envia a backup el archivo: {}".format(file))

            if os.path.exists(os.path.join(os.path.join(config.path_gen, folder), file)):
                os.remove(os.path.join(os.path.join(config.path_gen, folder), file))

            shutil.move(os.path.join(config.path_gen, file), os.path.join(config.path_gen, folder))


backup_files_gen()

dateTimeIni = datetime.now()

find_specific_row(sheet, dashboard, config.config_type_col)

try:
    os.remove(config.report)
except FileNotFoundError:
    print("Archivo de reporte no encontrado")
    logging.error("Archivo de reporte no encontrado")

wb.remove(sheet)
wb.save(config.report)

dateTimeFin = datetime.now()

print("Inicio Monitoreo: {}".format(dateTimeIni.strftime("%H:%M:%S")))
logging.info("Inicio Monitoreo: {}".format(dateTimeIni.strftime("%H:%M:%S")))
print("Fin Monitoreo: {}".format(dateTimeFin.strftime("%H:%M:%S")))
logging.info("Fin Monitoreo: {}".format(dateTimeFin.strftime("%H:%M:%S")))
ejecucion = dateTimeFin - dateTimeIni
print("Tiempo ejecución: {}".format(ejecucion))
logging.info("Tiempo ejecución: {}".format(ejecucion))

report_file = os.path.split(config.report)

email = Email(
        receiver = config.send_to,
        subject = config.subject,
        message = "Se ejecuto el reporte de monitoreo Diario de SAP, adjunto al correo esta el reporte. <br><br>" +
                  "<b>Inicio Monitoreo:</b> {} <br>".format(dateTimeIni.strftime("%H:%M:%S")) +
                  "<b>Fin Monitoreo:</b> {} <br><br>".format(dateTimeFin.strftime("%H:%M:%S")) +
                "<b>Tiempo ejecución:</b> {}".format(ejecucion),
        #file_to_attach = "result_SAP.xlsx")
        file_to_attach = config.report,
        image_path = config.path_gen)

if config.send_email == True:
    result_email = email.send_email()
    print(result_email)
    logging.info(result_email)

